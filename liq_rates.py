#!/usr/bin/python3
import mimetypes
import os
import smtplib
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart

import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup, Tag
from dotenv import load_dotenv
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv()

pd.options.mode.chained_assignment = None
pd_timestamp = pd.Timestamp
bs_tag = Tag
response_type = requests.models.Response


def get_url(url: str) -> response_type:
    response = requests.get(url)
    return response


def collect_data_for_df(res_text: str) -> list[list[str]]:
    all_rows = []
    soups = BeautifulSoup(res_text, "html.parser")
    table = soups.find("table", class_="data")
    trs = table.find_all("tr")  # type: ignore
    for tr in trs[1:]:
        date = tr.find_all("td")[0].text
        rate = tr.find_all("td")[1].text
        value = tr.find_all("td")[2].text
        all_rows.append([date, rate, value])
    return all_rows


def get_ruonia_df(dt_min: str, stop_date: str) -> pd.DataFrame:
    link = (
        f"https://cbr.ru/hd_base/ruonia/dynamics/"
        f"?UniDbQuery.Posted=True&"
        f"UniDbQuery.From={dt_min}&UniDbQuery.To={stop_date}"
    )
    res_text = get_url(link)
    all_rows = collect_data_for_df(res_text.text)
    df = pd.DataFrame(
        all_rows,
        columns=["Дата ставки", "Значение, %", "Объем сделок RUONIA, млрд руб."],
    )
    return df


def processing_ruonia(date_from: str, date_to: str) -> pd.DataFrame:
    ruonia = get_ruonia_df(date_from, date_to)
    ruonia["Дата ставки"] = pd.to_datetime(ruonia["Дата ставки"], format="%d.%m.%Y")
    ruonia["Значение, %"] = ruonia["Значение, %"].apply(
        lambda x: float(x.replace(",", "."))
    )
    return ruonia


def get_tables(
    response: response_type, category: str, date_from: str, date_to: str
) -> pd.DataFrame:
    soup = BeautifulSoup(response.text, "html.parser")
    headers = []
    if category == "ruonia":
        df = processing_ruonia(date_from, date_to)
    else:
        table = soup.find(
            "table",
            class_="table table-bordered table-condensed arch-table rrr matrix-table_",
        )

        for i in table.find_all("tr")[0]:  # type: ignore
            if isinstance(i, Tag):
                headers.append(i.text.strip())
        df = pd.DataFrame(columns=headers)
        for j in table.find_all("tr")[1:]:  # type: ignore
            row_data = j.find_all("td")
            row = [i.text for i in row_data]
            first_value = j.find("th").text.strip()
            row.insert(0, first_value)
            length = len(df)
            try:
                df.loc[length] = row
            except ValueError:
                row.append("")
                df.loc[length] = row
    return df


def convert_type_column(df: pd.DataFrame) -> pd.DataFrame:
    try:
        df["Дата ставки"] = pd.to_datetime(df["Дата ставки"], format="%d.%m.%Y")
    except (KeyError, ValueError, AttributeError, TypeError):
        df["Дата"] = pd.to_datetime(df["Дата"], format="%d.%m.%Y")
    for column in list(df.columns)[1:]:
        df[column] = df[column].apply(
            lambda x: (
                float(str(x.replace(",", ".").replace(" ", "")))
                if isinstance(x, str) and (x not in ["--", "—", " — "])
                else x
            )
        )
    return df


def write_to_excel(dfs: list[pd.DataFrame], categories: list[str], name_file: str) -> None:
    with pd.ExcelWriter(name_file) as writer:
        for i, df in enumerate(dfs):
            if categories[i] == "ruonia":
                need_columns = list(df.columns)[:3]
                df = df[need_columns]
            df = convert_type_column(df)
            df.sort_values(by=list(df.columns)[0], inplace=True)
            df.to_excel(
                writer, sheet_name=categories[i], index=False, freeze_panes=(1, 0)
            )


def processing_file(name_file: str) -> None:
    wb = openpyxl.load_workbook(name_file)
    sheetnames = wb.sheetnames
    for sheet in sheetnames:
        sheet_main = wb[sheet]
        max_column = sheet_main.max_column
        max_rows = sheet_main.max_row
        for k in range(1, max_column + 1):
            sheet_main.column_dimensions[get_column_letter(k)].width = 15
            for i in range(2, max_rows + 1):
                cell = sheet_main.cell(row=i, column=k)
                cell.alignment = Alignment(horizontal="right")
                cell_date = sheet_main.cell(row=i, column=1)
                cell_date.number_format = "DD.MM.YYYY"
        if sheet == "ruonia":
            sheet_main.column_dimensions[get_column_letter(3)].width = 25
    wb.move_sheet("roisfix implied", offset=-2)
    wb.save(name_file)


def get_df_with_key_rate(response: response_type) -> pd.DataFrame:
    df = pd.read_html(response.url, thousands=" ", flavor="html5lib")[0]
    return df


def processing_request(url: str, start: str, end: str) -> response_type:
    url = (
        f"{url}?UniDbQuery.Posted=True&UniDbQuery.From={start}&UniDbQuery.To={end}"
    )
    response = get_url(url)
    return response


def update_indicators(df: pd.DataFrame) -> pd.DataFrame:
    df.rename(columns={"Ставка": "CBR_key_rate"}, inplace=True)
    df["Дата"] = pd.to_datetime(df["Дата"], format="%d.%m.%Y")
    df["CBR_key_rate"] = df["CBR_key_rate"].apply(
        lambda x: 0.0 if pd.isnull(x) else float(x.replace(",", ".").replace(" ", ""))
    )
    return df


def get_key_rates(date_from: str, date_to: str) -> pd.DataFrame:
    url = "https://www.cbr.ru/hd_base/KeyRate/"
    response = processing_request(url, date_from, date_to)
    df = get_df_with_key_rate(response)
    df = update_indicators(df)
    return df


def division_by_hundred(df: pd.DataFrame, column: str) -> list[float]:
    rates = df[column].values.tolist()
    rates = [i / 100 for i in rates]
    return rates


def get_imputed_rates(rates_1: list[float], rates_2: list[float]) -> list[float]:
    imputed_rates = [
        round(100 * ((1 + rates_1[i]) ** 2 / (1 + rates_2[i]) - 1), 2)
        for i in range(len(rates_1))
    ]
    return imputed_rates


def get_rates(df: pd.DataFrame) -> tuple[list[float], ...]:
    rates_2y = division_by_hundred(df, "2Y")
    rates_1y = division_by_hundred(df, "1Y")
    rates_3m = division_by_hundred(df, "3M")
    rates_6m = division_by_hundred(df, "6M")
    rates_1m = division_by_hundred(df, "1M")
    rates_2m = division_by_hundred(df, "2M")

    rates_1y1y = get_imputed_rates(rates_2y, rates_1y)
    rates_3m3m = get_imputed_rates(rates_6m, rates_3m)
    rates_1m1m = get_imputed_rates(rates_2m, rates_1m)
    rates_6m6m = get_imputed_rates(rates_1y, rates_6m)
    return rates_1y1y, rates_3m3m, rates_1m1m, rates_6m6m


def count_spread(rates: list[float], key_rates: list[float]) -> list[float]:
    spread = [(rates[i] - key_rates[i]) * 100 for i in range(len(key_rates))]
    return spread


def get_spread_rates(
    key_rates: list[float],
    rates_1m1m: list[float],
    rates_3m3m: list[float],
    rates_1y1y: list[float],
    rates_6m6m: list[float],
    df: pd.DataFrame,
) -> pd.DataFrame:
    spread_1m1m = count_spread(rates_1m1m, key_rates)
    spread_3m3m = count_spread(rates_3m3m, key_rates)
    spread_1y1y = count_spread(rates_1y1y, key_rates)
    spread_6m6m = count_spread(rates_6m6m, key_rates)
    df[""] = ""
    df["spread_1m1m"] = spread_1m1m
    df["spread_3m3m"] = spread_3m3m
    df["spread_6m6m"] = spread_6m6m
    df["spread_1y1y"] = spread_1y1y
    return df


def write_to_excel_roisfix_implied(name_file: str, df: pd.DataFrame) -> None:
    with pd.ExcelWriter(
        name_file, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        df.to_excel(writer, index=False, sheet_name="roisfix implied")


def add_new_list(name_file: str, date_from: str, date_to: str) -> None:
    df_roisfix = pd.read_excel(name_file, sheet_name="roisfix")
    rates_1y1y, rates_3m3m, rates_1m1m, rates_6m6m = get_rates(df_roisfix)
    df = pd.DataFrame(columns=["Дата", "1M1M", "3M3M", "1Y1Y", "6M6M"])
    df = pd.concat([df_roisfix["Дата"]], axis=1)
    df["1M1M"] = rates_1m1m
    df["3M3M"] = rates_3m3m
    df["6M6M"] = rates_6m6m
    df["1Y1Y"] = rates_1y1y
    df[" "] = ""
    df_key_rates = get_key_rates(date_from, date_to)
    df_merge = df.merge(df_key_rates, how="left", on=["Дата"])
    key_rates = df_merge["CBR_key_rate"].values.tolist()
    df_merge = get_spread_rates(
        key_rates, rates_1m1m, rates_3m3m, rates_1y1y, rates_6m6m, df_merge
    )
    write_to_excel_roisfix_implied(name_file, df_merge)


def get_dataframes_from_categories(
    categories: list[str], date_from: str, date_to: str
) -> list[pd.DataFrame]:
    all_dfs = []
    for category in categories:
        link = f"http://{category}.ru/archive?date_from={date_from}&date_to={date_to}"
        response = get_url(link)
        df = get_tables(response, category, date_from, date_to)
        all_dfs.append(df)
    return all_dfs


def processing_dates(dates: list[bs_tag]) -> list[pd_timestamp]:
    months = {
        "января": "01",  "февраля": "02", "марта": "03", "апреля": "04",
        "мая": "05", "июня": "06", "июля": "07", "августа": "08",
        "сентября": "09", "октября": "10", "ноября": "11", "декабря": "12",
    }
    all_dates = []
    for day in dates:
        one_date = []
        try:
            titles = day.find_all("div", class_="title")
            titles = [title.text.strip().replace("\xa0", " ") for title in titles]  # type: ignore
            title = [
                title if "по ключевой ставке" in title else None for title in titles
            ][0]
            meeting_date = (
                day.find("div", class_="date col-md-5")  # type: ignore
                .text.strip()
                .replace("\xa0", " ")
                .replace(" года", "")
            )
            if len(meeting_date.split()) == 3:
                for month, month_new in months.items():
                    if month in meeting_date:
                        meeting_date = meeting_date.replace(month, month_new)
                meeting_date_new = datetime.strptime(meeting_date, "%d %m %Y")
                one_date.extend([meeting_date_new, title])
                all_dates.append(one_date)
        except IndexError:
            pass
    df_meeting = pd.DataFrame(all_dates, columns=["Date", "Title"])
    df_meeting = df_meeting.dropna(axis=0)
    df_meeting.index = [i for i in range(df_meeting.shape[0])]  # type: ignore
    return df_meeting["Date"].tolist()


def get_meeting_days() -> list[bs_tag]:
    url = "https://www.cbr.ru/dkp/cal_mp/#t11"
    r = get_url(url)
    soup = BeautifulSoup(r.text, "html.parser")
    dates = soup.find_all("div", class_="main-events_day")
    return dates


def select_meeting_dates(name_file: str, meeting_dates: list[pd_timestamp]) -> None:
    wb = openpyxl.load_workbook(name_file)
    wb_main = wb["roisfix implied"]
    max_column = wb_main.max_column
    roisfix_implied = pd.read_excel(name_file, sheet_name="roisfix implied")
    for i, day in enumerate(roisfix_implied["Дата"].tolist()):
        if day in meeting_dates:
            for j in range(1, max_column + 1):
                cell = wb_main.cell(row=i + 2, column=j)
                cell.fill = PatternFill("solid", fgColor="EDDFF6")
    wb.save(name_file)


def create_message(file_path: str) -> MIMEMultipart:
    message = MIMEMultipart()
    message["Subject"] = "LIQ Rates"
    ftype, _ = mimetypes.guess_type(file_path)
    file_type, subtype = ftype.split("/")  # type: ignore
    if file_type == "application":
        with open(file_path, "rb") as f:
            file = MIMEApplication(f.read(), subtype)

    file.add_header("content-disposition", "attachment", filename=file_path)
    message.attach(file)
    return message


def send_email(file_path: str) -> None:
    from_email = os.getenv("FROM_EMAIL")
    password = os.getenv("PASSWORD")

    smtp = smtplib.SMTP("smtp.mail.ru", 587)
    smtp.starttls()
    smtp.login(from_email, password)  # type: ignore

    message = create_message(file_path)
    smtp.sendmail(from_email, from_email, message.as_string())  # type: ignore
    smtp.quit()


def main():
    categories = ["ruonia", "roisfix", "nfeaswap", "rurepo"]
    date_from = "01.01.2024"
    date_to = datetime.now().date().strftime("%d.%m.%Y")
    name_file = "LIQ_Rates_Output.xlsx"
    all_dfs = get_dataframes_from_categories(categories, date_from, date_to)
    if any(df.empty is True for df in all_dfs):
        print("В списке есть пустой датафрейм")
    else:
        write_to_excel(all_dfs, categories, name_file)
        add_new_list(name_file, date_from, date_to)
    processing_file(name_file)
    meeting_dates = get_meeting_days()
    meeting_dates = processing_dates(meeting_dates)
    select_meeting_dates(name_file, meeting_dates)
    send_email(name_file)


if __name__ == "__main__":
    main()
